# bulk_enrollment_api.py
"""
Bulk device enrollment via CSV/XLSX
===================================
Super-admin (impersonating) or company-admin downloads a template file, fills
rows, and uploads. The flow is validate -> commit:

  GET  /bulk-devices/template.csv            download the CSV template
  GET  /bulk-devices/template.xlsx           download the XLSX template
  POST /bulk-devices/validate                (multipart) parse + validate; writes
                                             NOTHING; returns a job_id + preview
  POST /bulk-devices/commit                  {job_id}: create shops/groups/devices
  GET  /bulk-devices/jobs/{job_id}           job status + report
  GET  /bulk-devices/pending                 list pending (unclaimed) devices
  POST /bulk-devices/claim                   {device_id, mobile_id}: bind a real
                                             ANDROID_ID to a pending row

Rows WITH a device_id (the ANDROID_ID) auto-enroll the moment that physical
device polls GET /device/{id}/online (which returns 200 once the row exists).
Rows WITHOUT one become pending; an installer reads the ANDROID_ID shown on the
device's unenrolled screen and claims the row from the dashboard.

Tenant-safety (per the audit): tenant_id always comes from the session, never the
file; a device_id already used under ANOTHER tenant is rejected (global mobile_id
lookups ignore tenant); max_devices is enforced here (it is enforced nowhere else).
Content links (group_video copy) are intentionally NOT created here to avoid the
known tenant_id=1 default leak on device_video_shop_group — content flows through
the normal group pipeline / template playlist.
"""
import csv
import io
import json
import logging
import re
import secrets
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from database import pg_conn
from tenant_context import TenantContext, log_audit, require_tenant_context
import template_api as tpl_api

logger = logging.getLogger("bulk_enrollment_api")
router = APIRouter()


def _require_manage_devices(ctx: TenantContext):
    """Bulk device import creates/updates screens + per-screen content, so it
    needs the same authority as managing devices. The routes are directly
    callable, so this server-side check — not just a hidden button — is the
    real gate."""
    if not ctx.has_permission("manage_devices"):
        raise HTTPException(status_code=403, detail="Permission denied: manage_devices required")


def _bulk_mode(ctx: TenantContext) -> str:
    """'full' (manage_devices) = create/claim screens, assignments AND content.
    'content' (any template-content edit permission, i.e. the editor role) =
    content.* columns on EXISTING screens only — rows that would add screens or
    change name/location/group/template are validation errors, and the writes
    ride the template-content approval gate like the zone editor does."""
    if ctx.has_permission("manage_devices"):
        return "full"
    if any(ctx.has_permission(p) for p in tpl_api._CONTENT_EDIT_PERMS):
        return "content"
    raise HTTPException(status_code=403, detail="Permission denied: manage_devices required")

BASE_COLUMNS = ["device_name", "shop_name", "group_name", "template", "device_id", "resolution", "notes"]
BASE_WIDTHS = [26, 22, 18, 22, 22, 12, 30]
COLUMNS = BASE_COLUMNS  # back-compat alias
REQUIRED = ["device_name", "shop_name"]
RESOLUTION_RE = re.compile(r"^\d{2,5}x\d{2,5}$")
MOBILE_ID_RE = re.compile(r"^[A-Za-z0-9._:-]{1,64}$")
PENDING_PREFIX = "pending:"
MAX_ROWS = 20000
COMMIT_BATCH = 500

BASE_EXAMPLE = [
    ["Main Entrance Screen", "Shop Karachi 12", "North Region", "", "a1b2c3d4e5f6a7b8", "1080x1920", "landscape TV at door"],
    ["Checkout Screen", "Shop Karachi 12", "North Region", "", "", "", "device id unknown - will be claimed on site"],
]
BASE_INSTRUCTIONS = [
    "DIGIX bulk device import — fill one row per screen, then upload this file.",
    "device_name (required): a friendly name for the screen.",
    "shop_name (required): the shop this screen belongs to; created automatically if new.",
    "group_name (optional): a group for the screen; created automatically if new.",
    "template (optional): the screen template THIS screen should render, by exact name. Blank = inherit from its group / the company default. Templates are never created from this sheet — the name must match one linked to your company.",
    "device_id (optional): the device's ANDROID_ID. If you know it, the screen auto-enrolls when it powers on. Leave blank to create a 'pending' screen you claim on site.",
    "resolution (optional): e.g. 1080x1920. Auto-detected from the device if blank.",
    "notes (optional): free text, not shown on screens.",
]
# Per-zone example + guidance shown for the content columns (device-level content).
_ZONE_HELP = {
    "media": ("https://example.com/promo.jpg",
              "an image/video URL, an s3:// path, or the NAME of a video/advertisement already uploaded"),
    "qr": ("https://menu.example.com/shop12",
           "a LINK to turn into a QR code, OR an image URL / s3:// path / library name to show as-is"),
    "text": ("50% OFF this week", "the text to show on this screen"),
    "bg": ("#111827  (or  #0a1628 -> #f59e0b @135  for a gradient, or an image URL)",
           "a #hex color, a gradient like '#0a1628 -> #f59e0b @135', or an image URL/name"),
    "fit": ("contain",
            "how the image/video fits its box: cover (fill & crop), contain (show the whole thing), fill (stretch), or none. Blank = leave as-is. "
            "If contain makes a video look tiny, the FILE itself has black bars baked in — use cover, or re-export the file at the box's size."),
}


def content_columns_for(zones) -> list:
    """[(header, zone_key, field, zone_type)] for each tenant-fillable zone.

    Deliberately CONTENT-only (user decision 2026-07-17): the sheet carries
    text / media / QR values; backgrounds, colors and every other style live
    in the designer. Old sheets that still contain content.*.bg columns parse
    harmlessly — unknown headers are ignored.

    EVERY text/ticker box gets a column, whatever its binding (user decision
    2026-07-18): a designed or name-bound text is only the default, and a
    sheet value overrides it on that screen (see template_api.resolve_zone).
    """
    cols = []
    for z in zones or []:
        if not tpl_api.takes_tenant_content(z):
            continue
        key, zt = z.get("key"), z.get("type")
        if zt in ("text", "ticker"):
            cols.append((f"content.{key}.text", key, "text", zt))
        elif zt == "media":
            cols.append((f"content.{key}.media", key, "media", zt))
            cols.append((f"content.{key}.fit", key, "fit", zt))
        elif zt == "qr":
            cols.append((f"content.{key}.qr", key, "qr", zt))
    return cols


def _tenant_content_zones(cur, tenant_id: int):
    """(content_zones, default_template, all_usable_templates).

    Zones are the UNION across every template the company can use (default
    first; the first definition of a key wins) — content is keyed by zone_key,
    so one 'promo' column fills the promo box on every template that has it.
    """
    templates = tpl_api._company_template_choices(cur, tenant_id)
    if not templates:
        return [], None, []
    default = next((t for t in templates if t.get("is_default")), templates[0])
    seen: Dict[str, Dict] = {}
    for t in templates:
        for z in t["zones"]:
            if not tpl_api.takes_tenant_content(z):
                continue
            seen.setdefault(z.get("key"), z)
    return list(seen.values()), default, templates


def _library_names(cur, tenant_id: int) -> Dict[str, str]:
    """s3_link -> library item name (videos + advertisements). Exported cells
    prefer the NAME: it is what users recognize, and re-parsing a name restores
    media_type from the library row (an s3://-only cell would re-derive it from
    the file extension, which can disagree for extension-less objects)."""
    names: Dict[str, str] = {}
    cur.execute("SELECT s3_link, ad_name FROM public.advertisement WHERE tenant_id = %s;", (tenant_id,))
    names.update({r[0]: r[1] for r in cur.fetchall() if r[0] and r[1]})
    # Videos AFTER ads: resolve_media_value checks the video table first, so on a
    # name collision the video row wins the lookup — let it win the export too.
    cur.execute("SELECT s3_link, video_name FROM public.video WHERE tenant_id = %s;", (tenant_id,))
    names.update({r[0]: r[1] for r in cur.fetchall() if r[0] and r[1]})
    return names


def _media_cell(pl: Dict, lib_names: Dict[str, str]) -> str:
    """A media payload → sheet cell that parses back to the SAME payload."""
    name = lib_names.get(pl.get("media_s3") or "")
    if name:
        return name
    v = pl.get("media_url") or pl.get("media_s3") or ""
    if not v:
        return ""
    derived = "video" if tpl_api.VIDEO_EXT_RE.search(v) else "image"
    if pl.get("media_type") and pl["media_type"] != derived:
        # The cell grammar can't carry an explicit type (it re-derives from the
        # extension). Exporting this value would flip the type on re-upload —
        # leave it blank instead (blank = leave as-is; the read-only sheet still
        # shows what plays).
        return ""
    return v


def _cell_of(field: str, pl: Dict, lib_names: Dict[str, str]) -> str:
    """Inverse of _parse_row_content for one column: a cell value that re-parses
    to the stored payload, so re-uploading an unmodified export is a no-op."""
    if not pl:
        return ""
    if field == "text":
        return pl.get("text") or ""
    if field == "fit":
        return pl.get("fit_mode") or ""
    if field == "bg":
        g = pl.get("bg_gradient") or {}
        if g.get("stops"):
            return f"{' -> '.join(g['stops'])} @{g.get('angle', 135)}"
        return pl.get("bg_color") or pl.get("bg_image_url") or pl.get("bg_image_s3") or ""
    if field == "media":
        return _media_cell(pl, lib_names)
    if field == "qr":
        if pl.get("qr_link"):
            return pl["qr_link"]
        return _media_cell(pl, lib_names)
    return ""


def _fleet_rows(cur, tenant_id: int, ccols) -> Tuple[list, list]:
    """(sheet_rows, devices) — one row per existing screen (newest row per
    mobile_id, the row the player resolves). Content cells carry the screen's
    OWN device-scope values; blank = inherits group/location/company."""
    cur.execute("""
        SELECT DISTINCT ON (d.mobile_id)
               d.id, d.mobile_id, d.device_name,
               COALESCE(d.reported_resolution, d.resolution, ''),
               COALESCE(s.shop_name, ''), COALESCE(g.gname, ''), d.activation_code,
               COALESCE(st.name, ''), g.template_id
        FROM public.device d
        LEFT JOIN public.device_assignment da ON da.did = d.id
        LEFT JOIN public.shop s ON s.id = da.sid
        LEFT JOIN public."group" g ON g.id = da.gid
        LEFT JOIN public.screen_template st ON st.id = d.template_id
        WHERE d.tenant_id = %s
        ORDER BY d.mobile_id, d.id DESC;
    """, (tenant_id,))
    devs = sorted(cur.fetchall(), key=lambda r: (r[4], r[2] or ""))
    if not devs:
        return [], []
    lib_names = _library_names(cur, tenant_id) if ccols else {}
    own: Dict[int, Dict[str, Dict]] = {}
    if ccols:
        cur.execute("""
            SELECT device_id, zone_key, payload FROM public.template_zone_content
            WHERE tenant_id = %s AND scope = 'device' AND device_id = ANY(%s);
        """, (tenant_id, [d[0] for d in devs]))
        for did, zk, pl in cur.fetchall():
            own.setdefault(did, {})[zk] = json.loads(pl) if isinstance(pl, str) else pl
    rows = []
    for did, mid, name, res, shop, grp, activation, tpl_override, _grp_tpl in devs:
        pending = mid.startswith(PENDING_PREFIX)
        note = f"pending — claim on site (code {activation})" if pending and activation else ""
        # template cell = the screen's OWN override only; blank = inherits
        # (group/company) — mirrors the blank-means-leave-as-is grammar.
        base = [name or "", shop, grp, tpl_override, mid or "", res, note]
        rows.append(base + [_cell_of(f, (own.get(did) or {}).get(k) or {}, lib_names)
                            for (_h, k, f, _zt) in ccols])
    return rows, devs


_UNSET_FALLBACK = {
    # What a box with NO tenant content anywhere actually renders, by binding.
    "company.name": "(shows the company name)",
    "shop.name": "(shows the location name)",
    "device.name": "(shows the screen name)",
    "static": "(shows the template's designed text)",
}


def _effective_playback(cur, tenant_id: int, devs, ccols, templates=None,
                        zone_sources=None) -> list:
    """Read-only rows: what each screen actually renders per editable box right
    now, and which level set it (screen > group > location > company)."""
    zone_keys = list(dict.fromkeys(k for (_h, k, _f, _zt) in ccols))
    if not zone_keys or not devs:
        return []
    tpl_names = {t["id"]: t["name"] for t in (templates or [])}
    default_tpl = next((t["name"] for t in (templates or []) if t.get("is_default")), "")
    cur.execute("SELECT did, sid, gid FROM public.device_assignment WHERE did = ANY(%s);",
                ([d[0] for d in devs],))
    asg = {r[0]: (r[1], r[2]) for r in cur.fetchall()}
    company: Dict[str, Dict] = {}
    by_shop: Dict[int, Dict[str, Dict]] = {}
    by_dev: Dict[int, Dict[str, Dict]] = {}
    by_grp: Dict[int, Dict[str, Dict]] = {}
    cur.execute("SELECT zone_key, scope, shop_id, device_id, payload"
                " FROM public.template_zone_content WHERE tenant_id = %s;", (tenant_id,))
    for zk, scope, sid, did, pl in cur.fetchall():
        pl = json.loads(pl) if isinstance(pl, str) else pl
        if scope == "company":
            company[zk] = pl
        elif scope == "shop" and sid is not None:
            by_shop.setdefault(sid, {})[zk] = pl
        elif scope == "device" and did is not None:
            by_dev.setdefault(did, {})[zk] = pl
    cur.execute("SELECT zone_key, group_id, payload"
                " FROM public.template_zone_group_content WHERE tenant_id = %s;", (tenant_id,))
    for zk, gid, pl in cur.fetchall():
        by_grp.setdefault(gid, {})[zk] = json.loads(pl) if isinstance(pl, str) else pl
    out = []
    for did, mid, name, _res, shop, grp, _act, tpl_override, grp_tpl_id in devs:
        sid, gid = asg.get(did, (None, None))
        # Effective template: screen override > group override > company default.
        eff_tpl = tpl_override or tpl_names.get(grp_tpl_id, "") or default_tpl
        for zk in zone_keys:
            for level, store in (("this screen", by_dev.get(did)),
                                 ("its group", by_grp.get(gid) if gid else None),
                                 ("its location", by_shop.get(sid) if sid else None),
                                 ("company default", company)):
                pl = (store or {}).get(zk)
                if pl:
                    out.append([name or mid, shop, grp, eff_tpl, zk, level, _content_summary(pl)])
                    break
            else:
                fallback = _UNSET_FALLBACK.get((zone_sources or {}).get(zk))
                if fallback:
                    out.append([name or mid, shop, grp, eff_tpl, zk, "the template", fallback])
                else:
                    out.append([name or mid, shop, grp, eff_tpl, zk, "", "(nothing set — box is blank)"])
    return out


def _template_bundle(tenant_id: int):
    """Columns, current-fleet rows (or example rows for an empty fleet),
    read-only playback rows, and instructions for the tenant's linked template."""
    with pg_conn() as conn:
        with conn.cursor() as cur:
            content_cols, tpl, templates = _tenant_content_zones(cur, tenant_id)
            ccols = content_columns_for(content_cols)
            zone_sources = {z.get("key"): (z.get("binding") or {}).get("source", "static")
                            for z in content_cols}
            fleet, devs = _fleet_rows(cur, tenant_id, ccols)
            playback = _effective_playback(cur, tenant_id, devs, ccols, templates,
                                           zone_sources=zone_sources)
    headers = BASE_COLUMNS + [h for (h, _, _, _) in ccols]
    if fleet:
        rows = fleet
    else:
        rows = [list(r) + ["" for _ in ccols] for r in BASE_EXAMPLE]
        if ccols and rows:
            for i, (_h, _k, field, _zt) in enumerate(ccols):
                rows[0][len(BASE_COLUMNS) + i] = _ZONE_HELP.get(field, ("", ""))[0]
    instructions = list(BASE_INSTRUCTIONS)
    if fleet:
        instructions.insert(1, f"This file lists your CURRENT {len(fleet)} screen(s) — one row each. "
                               "Edit a value or append new rows, then upload. A blank cell always means "
                               "'leave as-is', so re-uploading an untouched file changes nothing.")
    if templates and len(templates) > 1:
        names = ", ".join(f"'{t['name']}' ({t['orientation']} {t['design_width']}×{t['design_height']})"
                          for t in templates)
        instructions.insert(2 if fleet else 1,
                            f"Templates available to the template column: {names}. "
                            f"Default (used when blank): '{tpl['name']}'.")
    if ccols:
        instructions.append("")
        instructions.append(f"Screen content (from the linked template '{tpl['name']}' v{tpl.get('version', 1)}"
                            + (f" plus {len(templates) - 1} more linked template(s); a column fills the box with "
                               f"that name on WHICHEVER template a screen renders" if len(templates) > 1 else "")
                            + ") — applies to THAT screen only and OVERRIDES its group/location/company content; "
                              "blank = keep inheriting:")
        for (h, _k, field, _zt) in ccols:
            instructions.append(f"{h}: {_ZONE_HELP.get(field, ('', 'content'))[1]}")
    else:
        instructions.append("")
        instructions.append("(No screen template is linked to this company yet, so there are no content columns.)")
    return headers, rows, instructions, ccols, playback, bool(fleet)


def _csv_bytes(headers, rows, instructions) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    for line in instructions:
        w.writerow(["# " + line] if line else [])
    w.writerow(headers)
    for row in rows:
        w.writerow(row)
    return buf.getvalue().encode("utf-8")


# ─────────────────────────── models ───────────────────────────

class CommitIn(BaseModel):
    job_id: int
    auto_create: bool = True  # create missing shops/groups automatically


class ClaimIn(BaseModel):
    device_id: int                      # the pending device row's id
    mobile_id: str = Field(min_length=1, max_length=64)  # the real ANDROID_ID


# ─────────────────────────── template download ───────────────────────────

@router.get("/bulk-devices/template.csv")
def template_csv(ctx: TenantContext = Depends(require_tenant_context)):
    _bulk_mode(ctx)  # content editors may download the sheet too
    headers, rows, instructions, _ccols, _playback, _is_fleet = _template_bundle(ctx.active_tenant_id)
    return StreamingResponse(
        io.BytesIO(_csv_bytes(headers, rows, instructions)), media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="digix-devices-template.csv"'},
    )


@router.get("/bulk-devices/template.xlsx")
def template_xlsx(ctx: TenantContext = Depends(require_tenant_context)):
    _bulk_mode(ctx)  # content editors may download the sheet too
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=501, detail="XLSX export unavailable; use the CSV template")
    headers, rows, instructions, _ccols, playback, is_fleet = _template_bundle(ctx.active_tenant_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"
    header_fill = PatternFill("solid", fgColor="0A1628")
    content_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for i, cell in enumerate(ws[1]):
        cell.fill = content_fill if i >= len(BASE_COLUMNS) else header_fill
        cell.font = header_font
    for row in rows:
        ws.append(row)
    for i in range(len(headers)):
        ws.column_dimensions[_col_letter(i + 1)].width = 30 if i >= len(BASE_COLUMNS) else BASE_WIDTHS[i]

    # A blank writable cell means "inherited" — which reads as missing data.
    # Annotate every blank content/template cell with WHAT it inherits and from
    # where (from the playback rows). Comments are ignored on upload, so the
    # round-trip stays a no-op.
    if is_fleet and playback:
        from openpyxl.comments import Comment
        inherited = {}   # (screen, zone_key) -> (level, showing)
        eff_tpl = {}     # screen -> effective template name
        for scr, _loc, _grp, tpl_name, zone, set_at, showing in playback:
            eff_tpl[scr] = tpl_name
            if set_at and set_at != "this screen":
                inherited[(scr, zone)] = (set_at, showing)
        tpl_col = BASE_COLUMNS.index("template") + 1
        for ri, r in enumerate(rows, start=2):  # row 1 = headers
            screen = r[0] or r[BASE_COLUMNS.index("device_id")]
            if not r[tpl_col - 1] and eff_tpl.get(screen):
                ws.cell(row=ri, column=tpl_col).comment = Comment(
                    f"Inherited — this screen renders '{eff_tpl[screen]}'.\n"
                    "Type a linked template's name to override just this screen.", "DIGIX")
            for ci, h in enumerate(headers):
                if not h.startswith("content.") or (ci < len(r) and r[ci]):
                    continue
                zone_key = h.split(".")[1]
                info = inherited.get((screen, zone_key))
                if info:
                    ws.cell(row=ri, column=ci + 1).comment = Comment(
                        f"Inherited from {info[0]}: {info[1]}\n"
                        "Type a value to override THIS screen only; blank keeps inheriting.", "DIGIX")
    if playback:
        # Visibility, not input: what every screen renders in each editable box
        # and which level set it. Only the first sheet is read on upload.
        pb = wb.create_sheet("Current playback (read-only)")
        pb.append(["screen", "location", "group", "template", "box", "set at", "currently showing"])
        for cell in pb[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in playback:
            pb.append(row)
        for i, width in enumerate([26, 22, 18, 22, 20, 18, 44]):
            pb.column_dimensions[_col_letter(i + 1)].width = width
    info = wb.create_sheet("Instructions")
    for line in instructions:
        info.append([line])
    info.column_dimensions["A"].width = 120
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="digix-devices-template.xlsx"'},
    )


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# ─────────────────────────── parsing ───────────────────────────

def _parse_upload(filename: str, data: bytes) -> List[Dict[str, str]]:
    """Return a list of {column: value} dicts. Skips comment/blank/header rows."""
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(data)
    return _parse_csv(data)


def _norm_header(cells: List[str]) -> Optional[List[str]]:
    lowered = [str(c or "").strip().lower() for c in cells]
    if "device_name" in lowered and "shop_name" in lowered:
        return lowered
    return None


def _rows_from(header: List[str], data_rows: List[List[str]]) -> List[Dict[str, str]]:
    base_idx = {col: header.index(col) for col in BASE_COLUMNS if col in header}
    # Any header of the form content.<zone>.<field> is a template content column.
    content_idx = {h: i for i, h in enumerate(header) if h.startswith("content.")}
    out = []
    for cells in data_rows:
        if not any(str(c or "").strip() for c in cells):
            continue

        def cell(i):
            return str(cells[i]).strip() if (i is not None and i < len(cells) and cells[i] is not None) else ""

        rec = {col: cell(base_idx.get(col)) for col in BASE_COLUMNS}
        rec["_content"] = {h: cell(i) for h, i in content_idx.items() if cell(i)}
        out.append(rec)
    return out


def _parse_csv(data: bytes) -> List[Dict[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))
    header = None
    body: List[List[str]] = []
    for cells in reader:
        if not cells:
            continue
        if cells[0].strip().startswith("#"):
            continue
        if header is None:
            h = _norm_header(cells)
            if h:
                header = h
                continue
            # first non-comment row isn't a header -> assume canonical order
            header = COLUMNS[:]
            body.append(cells)
            continue
        body.append(cells)
    if header is None:
        raise HTTPException(status_code=422, detail="No data rows found in the file")
    return _rows_from(header, body)


def _parse_xlsx(data: bytes) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=501, detail="XLSX upload unavailable; save as CSV and re-upload")
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
    header = None
    body = []
    for cells in rows:
        strs = [str(c).strip() for c in cells]
        if strs and strs[0].startswith("#"):
            continue
        if header is None:
            h = _norm_header(strs)
            if h:
                header = h
                continue
            continue
        body.append(strs)
    if header is None:
        raise HTTPException(status_code=422, detail="No header row (device_name, shop_name, …) found in the sheet")
    return _rows_from(header, body)


# ─────────────────────────── validation ───────────────────────────

def _parse_row_content(cur, tenant_id, content_cols, raw_content):
    """Parse a row's content cells → ({zone_key: payload}, [errors]). content_cols
    is [(header, zone_key, field, zone_type)]. Needs a DB cursor for name lookups."""
    payloads: Dict[str, Dict[str, Any]] = {}
    ztypes: Dict[str, str] = {}
    errors: List[str] = []
    for (header, key, field, zt) in content_cols:
        val = (raw_content or {}).get(header, "").strip()
        if not val:
            continue
        ztypes[key] = zt
        pl = payloads.setdefault(key, {})
        try:
            if field == "text":
                pl["text"] = val
            elif field == "bg":
                pl.update(tpl_api.parse_bg_value(cur, tenant_id, val))
            elif field == "media":
                pl.update(tpl_api.resolve_media_value(cur, tenant_id, val))
            elif field == "fit":
                # Sets only the media Fit; validate_content_payload rejects a bad
                # value. Written into fit_mode, which the players read (style.fit_mode)
                # — the same field the "Screen content" Fit dropdown controls.
                pl["fit_mode"] = val.lower()
            elif field == "qr":
                pl.update(tpl_api.parse_qr_value(cur, tenant_id, val))
        except ValueError as e:
            errors.append(f"{header}: {e}")
    for key, pl in payloads.items():
        for msg in tpl_api.validate_content_payload(ztypes[key], pl):
            errors.append(f"content.{key}: {msg}")
    return payloads, errors


def _content_summary(payload) -> str:
    """Short human-readable value for a content payload (for the change preview)."""
    if not payload:
        return ""
    if payload.get("text"):
        return (payload["text"][:40])
    if payload.get("qr_link"):
        return payload["qr_link"]
    if payload.get("media_url"):
        return payload["media_url"].split("?")[0].split("/")[-1]
    if payload.get("media_s3"):
        return payload["media_s3"].split("/")[-1]
    # A style-only payload (e.g. just Fit) — describe it, don't imply media set.
    if payload.get("fit_mode"):
        return f"fit: {payload['fit_mode']}"
    return "set"


def _current_states(cur, tenant_id: int, mobile_ids) -> Dict[str, Dict[str, Any]]:
    """{mobile_id: {id, name, shop, group, content:{zone_key: payload}}} for existing
    devices in this tenant — the baseline a re-upload is diffed against."""
    out: Dict[str, Dict[str, Any]] = {}
    ids = [m for m in (mobile_ids or set()) if m]
    if not ids:
        return out
    # id ASC so that with duplicate mobile_ids (stale ghost rows) the NEWEST row
    # ends up in out[mid] — the same row the player resolves and the commit writes.
    cur.execute("""
        SELECT d.id, d.mobile_id, d.device_name, s.shop_name, g.gname, st.name
        FROM public.device d
        LEFT JOIN public.device_assignment da ON da.did = d.id
        LEFT JOIN public.shop s ON s.id = da.sid
        LEFT JOIN public."group" g ON g.id = da.gid
        LEFT JOIN public.screen_template st ON st.id = d.template_id
        WHERE d.tenant_id = %s AND d.mobile_id = ANY(%s)
        ORDER BY d.id ASC;
    """, (tenant_id, ids))
    did_to_mid: Dict[int, str] = {}
    for did, mid, name, shop, grp, tpl_name in cur.fetchall():
        stale = out.get(mid)
        if stale:  # older duplicate loses its reverse mapping too
            did_to_mid.pop(stale["id"], None)
        out[mid] = {"id": did, "name": name or "", "shop": shop or "", "group": grp or "",
                    "template": tpl_name or "", "content": {}}
        did_to_mid[did] = mid
    if did_to_mid:
        cur.execute("""
            SELECT device_id, zone_key, payload FROM public.template_zone_content
            WHERE tenant_id = %s AND scope = 'device' AND device_id = ANY(%s);
        """, (tenant_id, list(did_to_mid.keys())))
        for did, zk, payload in cur.fetchall():
            if isinstance(payload, str):
                payload = json.loads(payload)
            mid = did_to_mid.get(did)
            if mid:
                out[mid]["content"][zk] = payload
    return out


def _row_changes(device_name: str, shop_name: str, group_name: str,
                 row_content: Dict, current: Dict, template_name: str = "") -> List[Dict[str, str]]:
    """What a re-upload row changes on an EXISTING device: [{field, from, to}].
    A blank cell means 'leave as-is' — only a non-blank value that differs counts."""
    changes: List[Dict[str, str]] = []

    def diff(field, new_val, cur_val):
        if new_val and new_val != (cur_val or ""):
            changes.append({"field": field, "from": cur_val or "", "to": new_val})

    diff("name", device_name, current.get("name"))
    diff("location", shop_name, current.get("shop"))
    diff("group", group_name, current.get("group"))
    diff("template", template_name, current.get("template"))
    cur_content = current.get("content") or {}
    for zk, payload in (row_content or {}).items():
        cur_p = cur_content.get(zk) or {}
        # Subset comparison: only the fields the sheet sets. The stored payload may
        # carry extra server-added keys (e.g. qr_generated_s3) that the sheet never
        # provides — a full-dict compare would flag every QR/media as changed.
        if payload and any(cur_p.get(k) != v for k, v in payload.items()):
            changes.append({"field": f"content.{zk}",
                            "from": _content_summary(cur_p),
                            "to": _content_summary(payload)})
    return changes


def _normalize_name(v: str) -> str:
    return " ".join((v or "").split()).casefold()


def _name_guard(raw: str, existing: Optional[Dict[str, str]], kind: str) -> Optional[str]:
    """None if `raw` is an exact existing name or a genuinely NEW one; an error
    message when it near-misses an existing name (case/spacing) — those are
    typos that would silently auto-create a junk duplicate."""
    if not raw or existing is None:
        return None
    canonical = existing.get(_normalize_name(raw))
    if canonical is not None and canonical != raw:
        return f"{kind} '{raw}' doesn't match any existing {kind} exactly — use '{canonical}'"
    return None


def validate_rows(rows: List[Dict[str, str]], existing_device_count: int, max_devices: int,
                  mobile_ids_this_tenant: set, mobile_ids_other_tenant: set,
                  content_cols=None, cur=None, tenant_id=None,
                  existing_shops: Optional[Dict[str, str]] = None,
                  existing_groups: Optional[Dict[str, str]] = None,
                  template_names: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    """Pure validation. Returns preview {rows, errors, summary}. Writes nothing."""
    errors: List[Dict[str, Any]] = []
    seen_ids: Dict[str, int] = {}
    shops, groups = set(), set()
    valid_new = 0        # devices that will be created
    with_id = 0
    pending = 0
    skipped = 0          # already exist in this tenant (idempotent re-run)
    updated = 0          # existing devices with field/content changes
    unchanged = 0        # existing devices the sheet leaves as-is
    # Baseline for the change preview (existing devices only).
    current_states = (_current_states(cur, tenant_id, mobile_ids_this_tenant)
                      if (cur is not None and tenant_id is not None) else {})

    normalized: List[Dict[str, Any]] = []
    for i, rec in enumerate(rows):
        rownum = i + 1
        row_errors = []
        dev_id = rec.get("device_id", "").strip()
        # Required fields gate NEW rows only. For a screen that already exists
        # (fleet-export round trip) a blank cell means "leave as-is" — a screen
        # with no location yet must re-upload cleanly, not fail "shop_name is
        # required".
        if dev_id not in mobile_ids_this_tenant:
            for col in REQUIRED:
                if not rec.get(col):
                    row_errors.append(f"{col} is required")
        for raw, existing, kind in ((rec.get("shop_name", "").strip(), existing_shops, "location"),
                                    (rec.get("group_name", "").strip(), existing_groups, "group")):
            guard = _name_guard(raw, existing, kind)
            if guard:
                row_errors.append(guard)
        # Templates are NEVER created from the sheet — the cell must name one
        # of the company's linked templates exactly (blank = inherit).
        tpl_cell = rec.get("template", "").strip()
        if tpl_cell and template_names is not None and tpl_cell not in template_names.values():
            canonical = template_names.get(_normalize_name(tpl_cell))
            if canonical:
                row_errors.append(f"template '{tpl_cell}' doesn't match exactly — use '{canonical}'")
            else:
                avail = ", ".join(sorted(template_names.values())) or "none linked yet"
                row_errors.append(f"template '{tpl_cell}' isn't linked to your company (available: {avail})")
        if dev_id:
            if not MOBILE_ID_RE.match(dev_id):
                row_errors.append("device_id has invalid characters")
            elif dev_id in seen_ids:
                row_errors.append(f"duplicate device_id (also on row {seen_ids[dev_id]})")
            elif dev_id in mobile_ids_other_tenant:
                row_errors.append("device_id already belongs to another company")
            if (dev_id and dev_id.startswith(PENDING_PREFIX)
                    and dev_id not in mobile_ids_this_tenant):
                # A pending id that IS ours round-trips from the fleet export —
                # it flows through the normal update/unchanged path above.
                row_errors.append("device_id must be a real device id, not a placeholder")
        res = rec.get("resolution", "").strip()
        if res and not RESOLUTION_RE.match(res):
            row_errors.append("resolution must look like 1080x1920")

        # Template content cells (only when the company has a linked template).
        row_content = {}
        if content_cols and cur is not None:
            row_content, content_errs = _parse_row_content(cur, tenant_id, content_cols, rec.get("_content"))
            row_errors.extend(content_errs)

        will = "error"
        row_changes: List[Dict[str, str]] = []
        if not row_errors:
            if dev_id:
                seen_ids[dev_id] = rownum
                if dev_id in mobile_ids_this_tenant:
                    # Existing screen — diff the row against its current state so the
                    # operator sees exactly what a re-upload will change (and only that).
                    row_changes = _row_changes(
                        rec.get("device_name", "").strip(), rec.get("shop_name", "").strip(),
                        rec.get("group_name", "").strip(), row_content, current_states.get(dev_id, {}),
                        template_name=rec.get("template", "").strip())
                    if row_changes:
                        will = "update"
                        updated += 1
                    else:
                        will = "unchanged"
                        unchanged += 1
                else:
                    will = "create"
                    valid_new += 1
                    with_id += 1
            else:
                will = "pending"
                valid_new += 1
                pending += 1
            # Only genuinely NEW names — the preview's "will be created" list.
            if rec.get("shop_name") and (existing_shops is None
                                         or _normalize_name(rec["shop_name"]) not in existing_shops):
                shops.add(rec["shop_name"])
            if rec.get("group_name") and (existing_groups is None
                                          or _normalize_name(rec["group_name"]) not in existing_groups):
                groups.add(rec["group_name"])
        else:
            for e in row_errors:
                errors.append({"row": rownum, "reason": e})

        normalized.append({
            "row": rownum,
            "device_name": rec.get("device_name", ""),
            "shop_name": rec.get("shop_name", ""),
            "group_name": rec.get("group_name", ""),
            "template": rec.get("template", "").strip(),
            "device_id": dev_id,
            "resolution": res,
            "action": will,
            "content": row_content,
            "changes": row_changes,
        })

    after = existing_device_count + valid_new
    quota_ok = max_devices <= 0 or after <= max_devices
    if not quota_ok:
        errors.append({"row": 0, "reason": f"Quota exceeded: {existing_device_count} existing + {valid_new} new = {after} > limit {max_devices}"})

    return {
        "rows": normalized,
        "errors": errors,
        "summary": {
            "total_rows": len(rows),
            "will_create": with_id,
            "will_pending": pending,
            "will_skip": skipped,
            "will_update": updated,
            "will_unchanged": unchanged,
            "error_rows": len({e["row"] for e in errors if e["row"] > 0}),
            "new_shops": sorted(shops),
            "new_groups": sorted(groups),
            "content_columns": [h for (h, _, _, _) in (content_cols or [])],
            "quota": {"existing": existing_device_count, "new": valid_new, "after": after,
                      "max": max_devices, "ok": quota_ok},
            "valid": len(errors) == 0,
        },
    }


# ─────────────────────────── endpoints ───────────────────────────

def _restrict_preview_to_content(preview: Dict) -> Dict:
    """Content-only bulk (the editor role): rows that would ADD screens or change
    anything besides content.* columns become row errors, so the preview tells
    the editor exactly which cells their role can't apply. Nothing is created,
    so the new-shops/new-groups lists empty out."""
    for row in preview["rows"]:
        reasons = []
        if row["action"] in ("create", "pending"):
            reasons.append("Your role can only update content on existing screens — this row would add a screen")
        elif row["action"] == "update":
            noncontent = sorted({str(c.get("field", "")) for c in (row.get("changes") or [])
                                 if not str(c.get("field", "")).startswith("content.")})
            if noncontent:
                reasons.append("Your role can only change content columns — this row also changes: "
                               + ", ".join(noncontent))
        if reasons:
            row["action"] = "error"
            for msg in reasons:
                preview["errors"].append({"row": row["row"], "reason": msg})
    s = preview["summary"]
    s["will_create"] = 0
    s["will_pending"] = 0
    s["will_update"] = sum(1 for r in preview["rows"] if r["action"] == "update")
    s["new_shops"] = []
    s["new_groups"] = []
    s["error_rows"] = len({e["row"] for e in preview["errors"] if e["row"] > 0})
    s["valid"] = len(preview["errors"]) == 0
    s["content_only"] = True
    return preview


@router.post("/bulk-devices/validate")
async def bulk_validate(file: UploadFile = File(...), ctx: TenantContext = Depends(require_tenant_context)):
    mode = _bulk_mode(ctx)
    tenant_id = ctx.active_tenant_id
    data = await file.read()
    if not data:
        raise HTTPException(status_code=422, detail="Empty file")
    rows = _parse_upload(file.filename, data)
    if not rows:
        raise HTTPException(status_code=422, detail="No device rows found")
    if len(rows) > MAX_ROWS:
        raise HTTPException(status_code=422, detail=f"Too many rows ({len(rows)}); max {MAX_ROWS} per import")

    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM public.device WHERE tenant_id = %s;", (tenant_id,))
            existing = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(max_devices, 0) FROM public.company WHERE id = %s;", (tenant_id,))
            row = cur.fetchone()
            max_devices = row[0] if row else 0
            # Which device_ids in the file already exist, and in which tenant?
            ids = [r["device_id"].strip() for r in rows if r.get("device_id", "").strip()]
            mine, other = set(), set()
            if ids:
                cur.execute("SELECT mobile_id, tenant_id FROM public.device WHERE mobile_id = ANY(%s);", (ids,))
                for mid, tid in cur.fetchall():
                    (mine if tid == tenant_id else other).add(mid)

            # Content columns come from the company's usable templates (if any).
            content_zones, _tpl, _templates = _tenant_content_zones(cur, tenant_id)
            content_cols = content_columns_for(content_zones)
            # Template names the `template` column may reference (never created).
            template_names = {_normalize_name(t["name"]): t["name"] for t in _templates}

            # Existing location/group names (normalized -> canonical) so a
            # case/spacing near-miss errors with the real name instead of
            # silently auto-creating a duplicate.
            cur.execute("SELECT shop_name FROM public.shop WHERE tenant_id = %s;", (tenant_id,))
            existing_shops = {_normalize_name(r[0]): r[0] for r in cur.fetchall() if r[0]}
            cur.execute('SELECT gname FROM public."group" WHERE tenant_id = %s;', (tenant_id,))
            existing_groups = {_normalize_name(r[0]): r[0] for r in cur.fetchall() if r[0]}

            preview = validate_rows(rows, existing, max_devices, mine, other,
                                    content_cols=content_cols, cur=cur, tenant_id=tenant_id,
                                    existing_shops=existing_shops, existing_groups=existing_groups,
                                    template_names=template_names)
            if mode == "content":
                preview = _restrict_preview_to_content(preview)

            cur.execute("""
                INSERT INTO public.bulk_import_job
                    (tenant_id, uploaded_by, filename, status, total_rows, error_rows, payload, report)
                VALUES (%s, %s, %s, 'validated', %s, %s, %s::jsonb, %s::jsonb)
                RETURNING id;
            """, (tenant_id, ctx.user_id, file.filename, preview["summary"]["total_rows"],
                  preview["summary"]["error_rows"],
                  _json(preview["rows"]), _json(preview["summary"])))
            job_id = cur.fetchone()[0]
        conn.commit()
    return {"job_id": job_id, **preview}


def _commit_content_only(body: "CommitIn", ctx: TenantContext,
                         background_tasks: BackgroundTasks) -> Dict:
    """The editor role's bulk commit: apply ONLY content.* changes on existing
    screens (never creates screens/locations/groups, never renames/re-assigns/
    re-templates — validation already errored such rows). When the company's
    approval requirement applies to this user, the whole batch is parked as ONE
    content_change_request (request_type 'template_content_bulk') and nothing
    touches the live tables until a manager/admin approves."""
    tenant_id = ctx.active_tenant_id
    with pg_conn() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status, payload, report FROM public.bulk_import_job
                    WHERE id = %s AND tenant_id = %s;
                """, (body.job_id, tenant_id))
                job = cur.fetchone()
                if not job:
                    raise HTTPException(status_code=404, detail="Import job not found")
                if job[0] in ("committed", "pending_approval"):
                    raise HTTPException(status_code=409, detail="This import was already submitted")
                rows = job[1]
                summary = job[2] if isinstance(job[2], dict) else {}
                if not summary.get("valid", False):
                    raise HTTPException(status_code=422, detail="This import has validation errors — fix and re-upload")

                content_zones, _tpl, _templates = _tenant_content_zones(cur, tenant_id)
                zone_map = {z["key"]: z for z in content_zones}

                # Collect the merged per-screen zone payloads (same effective-content
                # seeding as the full bulk path, so a fit-only cell can't blank
                # inherited media). Merging at SUBMIT time makes the approval preview
                # show exactly what will land.
                items: List[Dict] = []
                devices_touched = set()
                for r in rows:
                    if r.get("action") not in ("update", "unchanged", "skip"):
                        continue
                    mid = (r.get("device_id") or "").strip()
                    changed_zones = {c["field"].split(".", 1)[1] for c in (r.get("changes") or [])
                                     if str(c.get("field", "")).startswith("content.")}
                    if not mid or not changed_zones:
                        continue
                    cur.execute("SELECT id FROM public.device WHERE tenant_id = %s AND mobile_id = %s"
                                " ORDER BY id DESC LIMIT 1;", (tenant_id, mid))
                    hit = cur.fetchone()
                    if not hit:
                        continue
                    did = hit[0]
                    sid, gid = tpl_api._device_shop_group_ids(cur, did)
                    effective = tpl_api._collapse_content(cur, tenant_id, sid, did, gid)
                    for zone_key in sorted(changed_zones):
                        payload = (r.get("content") or {}).get(zone_key)
                        zone = zone_map.get(zone_key)
                        if not zone or not payload:
                            continue
                        merged = dict(effective.get(zone_key) or {})
                        merged.pop("qr_generated_s3", None)
                        merged.update(dict(payload))
                        # A library/s3 pick from the sheet REPLACES the media — an
                        # inherited external media_url must not survive the merge
                        # (resolve_zone prefers media_url over media_s3, so the
                        # screen would keep playing the old URL).
                        if payload.get("media_s3") and not payload.get("media_url"):
                            merged.pop("media_url", None)
                        items.append({"device_id": did, "device_name": r.get("device_name") or mid,
                                      "zone_key": zone_key,
                                      "zone_label": zone.get("name") or zone_key,
                                      "payload": merged})
                        devices_touched.add(did)

                if items and tpl_api._approval_required(cur, ctx):
                    # Supersede this user's earlier pending bulk batch — a re-upload
                    # replaces it, same iterate-before-review semantics as the editor.
                    cur.execute("""
                        UPDATE public.content_change_request
                        SET status = 'cancelled', updated_at = NOW()
                        WHERE tenant_id = %s AND request_type = 'template_content_bulk'
                          AND status = 'pending' AND requested_by = %s
                        RETURNING id;
                    """, (tenant_id, ctx.user_id))
                    superseded = [x[0] for x in cur.fetchall()]
                    change_data = {"items": items, "job_id": body.job_id,
                                   "requested_by": ctx.user_id}
                    cur.execute("""
                        INSERT INTO public.content_change_request
                            (tenant_id, request_type, target_type, target_id, target_name,
                             change_data, requested_by, status, expires_at)
                        VALUES (%s, 'template_content_bulk', 'company', %s, %s, %s::jsonb, %s,
                                'pending', NOW() + INTERVAL '72 hours')
                        RETURNING id;
                    """, (tenant_id, tenant_id,
                          f"Bulk sheet — {len(devices_touched)} screen{'s' if len(devices_touched) != 1 else ''}",
                          _json(change_data), ctx.user_id))
                    request_id = cur.fetchone()[0]
                    report = {"status": "pending_approval", "request_id": request_id,
                              "screens": len(devices_touched), "content_changes": len(items),
                              "superseded_request_ids": superseded}
                    cur.execute("""
                        UPDATE public.bulk_import_job
                        SET status = 'pending_approval', report = %s::jsonb, finished_at = NOW()
                        WHERE id = %s;
                    """, (_json(report), body.job_id))
                    log_audit(conn, tenant_id, ctx.user_id, "template.content.bulk_approval_requested",
                              "content_change_request", request_id,
                              details={"job_id": body.job_id, "screens": len(devices_touched),
                                       "content_changes": len(items)})
                    tpl_api._notify_pending_approvals_count(background_tasks, cur, tenant_id)
                    conn.commit()
                    return {"job_id": body.job_id, "committed": False, **report}

                # Approver (or the gate is off): apply directly, content only.
                for it in items:
                    zone = zone_map[it["zone_key"]]
                    tpl_api._upsert_zone_content(conn, cur, tenant_id, zone, it["zone_key"],
                                                 "device", None, it["device_id"], it["payload"],
                                                 ctx.user_id)
                report = {"created": 0, "pending": 0, "skipped": 0,
                          "content_updated_existing": len(devices_touched),
                          "content_written": len(items), "shops": 0, "groups": 0}
                cur.execute("""
                    UPDATE public.bulk_import_job
                    SET status = 'committed', created_rows = 0, skipped_rows = 0,
                        report = %s::jsonb, finished_at = NOW()
                    WHERE id = %s;
                """, (_json(report), body.job_id))
                log_audit(conn, tenant_id, ctx.user_id, "devices.bulk_import", "device", None,
                          details=report)
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:
            conn.rollback()
            logger.error("content-only bulk commit failed (tenant %s job %s): %s",
                         tenant_id, body.job_id, e)
            raise HTTPException(status_code=500, detail="Bulk content update failed; nothing was written")
    return {"job_id": body.job_id, "committed": True, **report}


@router.post("/bulk-devices/commit")
def bulk_commit(body: CommitIn, background_tasks: BackgroundTasks,
                ctx: TenantContext = Depends(require_tenant_context)):
    if _bulk_mode(ctx) == "content":
        return _commit_content_only(body, ctx, background_tasks)
    tenant_id = ctx.active_tenant_id
    with pg_conn() as conn:
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status, payload, report FROM public.bulk_import_job
                    WHERE id = %s AND tenant_id = %s;
                """, (body.job_id, tenant_id))
                job = cur.fetchone()
                if not job:
                    raise HTTPException(status_code=404, detail="Import job not found")
                if job[0] == "committed":
                    raise HTTPException(status_code=409, detail="This import was already committed")
                rows = job[1]
                summary = job[2] if isinstance(job[2], dict) else {}
                if not summary.get("valid", False):
                    raise HTTPException(status_code=422, detail="This import has validation errors — fix and re-upload")

                # Re-check quota at commit time (another import may have run since).
                cur.execute("SELECT COUNT(*) FROM public.device WHERE tenant_id = %s;", (tenant_id,))
                existing = cur.fetchone()[0]
                cur.execute("SELECT COALESCE(max_devices, 0) FROM public.company WHERE id = %s;", (tenant_id,))
                max_devices = (cur.fetchone() or [0])[0]
                will_add = sum(1 for r in rows if r["action"] in ("create", "pending"))
                if max_devices > 0 and existing + will_add > max_devices:
                    raise HTTPException(status_code=409, detail=f"Quota exceeded at commit: {existing}+{will_add} > {max_devices}")

                # 1) upsert shops + groups, build name->id maps
                shop_ids = _upsert_named(cur, "shop", "shop_name",
                                         {r["shop_name"] for r in rows if r["shop_name"]}, tenant_id)
                group_ids = _upsert_named(cur, '"group"', "gname",
                                          {r["group_name"] for r in rows if r["group_name"]}, tenant_id)

                # Content zones of the company's usable templates, for writing
                # per-device content; template names for the `template` column.
                content_zones, _tpl, _templates = _tenant_content_zones(cur, tenant_id)
                zone_map = {z["key"]: z for z in content_zones}
                template_ids_by_name = {t["name"]: t["id"] for t in _templates}

                created = skipped = pending = updated = 0
                content_written = 0

                def write_row_content(did, row, only_zones=None):
                    """Per-device template content (device-scope override). only_zones
                    limits the write to the zones the change-preview flagged, so a
                    re-upload doesn't rewrite (and re-stamp) unchanged content."""
                    nonlocal content_written
                    # Resolve the device's location + group so a PARTIAL content write
                    # layers onto the EFFECTIVE content (device > group > location >
                    # company), not just the device-scope row. Without this, setting only
                    # a sub-field (e.g. .fit) on a screen that INHERITS its media would
                    # write a media-less device override that shadows the inherited media
                    # and blanks the zone. Seeding from the effective content keeps the
                    # shown media and applies the edit on top (a per-screen override, which
                    # is exactly what bulk device content is).
                    _sid, _gid = tpl_api._device_shop_group_ids(cur, did)
                    effective = tpl_api._collapse_content(cur, tenant_id, _sid, did, _gid)
                    for zone_key, payload in (row.get("content") or {}).items():
                        if only_zones is not None and zone_key not in only_zones:
                            continue
                        zone = zone_map.get(zone_key)
                        if not zone or not payload:
                            continue
                        # Full replace at write time, so seed from the effective payload
                        # and layer the sheet's sub-fields on top — never dropping a
                        # blank-left sibling (.bg), a UI-only field (text_color/fit_mode)
                        # or inherited media. Drop the server-managed QR image (regenerated).
                        merged = dict(effective.get(zone_key) or {})
                        merged.pop("qr_generated_s3", None)
                        merged.update(dict(payload))
                        # A library/s3 pick from the sheet REPLACES the media — an
                        # inherited external media_url must not survive the merge
                        # (resolve_zone prefers media_url over media_s3, so the
                        # screen would keep playing the old URL).
                        if payload.get("media_s3") and not payload.get("media_url"):
                            merged.pop("media_url", None)
                        tpl_api._upsert_zone_content(conn, cur, tenant_id, zone, zone_key,
                                                     "device", None, did, merged, ctx.user_id)
                        content_written += 1

                # 2) create devices row-by-row (safe + idempotent); assignment per device
                for r in rows:
                    if r["action"] == "error":
                        continue
                    if r["action"] in ("update", "unchanged", "skip"):
                        # Existing screen in this tenant. Apply ONLY the fields the
                        # change-preview flagged (name / location / group / content) —
                        # a blank cell leaves that value as-is, and unchanged rows are
                        # a no-op. (This replaces the old content-only "skip".)
                        mid = r.get("device_id")
                        chg = {c["field"] for c in (r.get("changes") or [])}
                        if not mid or not chg:
                            skipped += 1
                            continue
                        # Newest row per mobile_id — the SAME row the player resolves
                        # (device_template orders id DESC). Without this, a stale
                        # duplicate row could receive the content while the screen
                        # keeps rendering from the newest row: "diff showed but the
                        # device never updated".
                        cur.execute("SELECT id FROM public.device WHERE tenant_id = %s AND mobile_id = %s"
                                    " ORDER BY id DESC LIMIT 1;",
                                    (tenant_id, mid))
                        hit = cur.fetchone()
                        if not hit:
                            skipped += 1
                            continue
                        did = hit[0]
                        if "name" in chg and r.get("device_name"):
                            cur.execute("UPDATE public.device SET device_name = %s WHERE id = %s;",
                                        (r["device_name"], did))
                        if "template" in chg and r.get("template") in template_ids_by_name:
                            cur.execute("UPDATE public.device SET template_id = %s WHERE id = %s;",
                                        (template_ids_by_name[r["template"]], did))
                        loc_c, grp_c = ("location" in chg), ("group" in chg)
                        if loc_c or grp_c:
                            gid = group_ids.get(r["group_name"]) if r.get("group_name") else None
                            sid = shop_ids.get(r["shop_name"]) if r.get("shop_name") else None
                            cur.execute("""
                                INSERT INTO public.device_assignment (did, gid, sid, tenant_id)
                                VALUES (%s, %s, %s, %s)
                                ON CONFLICT (did) DO UPDATE SET
                                    gid = CASE WHEN %s THEN EXCLUDED.gid ELSE public.device_assignment.gid END,
                                    sid = CASE WHEN %s THEN EXCLUDED.sid ELSE public.device_assignment.sid END,
                                    updated_at = NOW();
                            """, (did, gid, sid, tenant_id, grp_c, loc_c))
                        changed_zones = {c["field"].split(".", 1)[1] for c in (r.get("changes") or [])
                                         if c["field"].startswith("content.")}
                        if changed_zones:
                            write_row_content(did, r, only_zones=changed_zones)
                        updated += 1
                        continue
                    if r["action"] == "pending":
                        code = _pending_code()
                        mobile_id = PENDING_PREFIX + code
                        activation = code
                    else:  # create
                        mobile_id = r["device_id"]
                        activation = None
                        # Guard: it may exist by now (idempotent re-run) — still
                        # refresh its content from the sheet (newest row, like the player).
                        cur.execute("SELECT id FROM public.device WHERE tenant_id = %s AND mobile_id = %s"
                                    " ORDER BY id DESC LIMIT 1;",
                                    (tenant_id, mobile_id))
                        existing_row = cur.fetchone()
                        if existing_row:
                            skipped += 1
                            if r.get("content"):
                                write_row_content(existing_row[0], r)
                                updated += 1
                            continue
                    cur.execute("""
                        INSERT INTO public.device
                            (mobile_id, download_status, is_online, is_active, last_online_at,
                             resolution, device_name, tenant_id, activation_code, template_id)
                        VALUES (%s, FALSE, FALSE, TRUE, NOW(), %s, %s, %s, %s, %s)
                        RETURNING id;
                    """, (mobile_id, r["resolution"] or None, r["device_name"] or None, tenant_id,
                          activation, template_ids_by_name.get(r.get("template") or "")))
                    did = cur.fetchone()[0]
                    gid = group_ids.get(r["group_name"]) if r["group_name"] else None
                    sid = shop_ids.get(r["shop_name"]) if r["shop_name"] else None
                    if sid or gid:
                        cur.execute("""
                            INSERT INTO public.device_assignment (did, gid, sid, tenant_id)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (did) DO UPDATE SET gid = EXCLUDED.gid, sid = EXCLUDED.sid, updated_at = NOW();
                        """, (did, gid, sid, tenant_id))
                    write_row_content(did, r)

                    if r["action"] == "pending":
                        pending += 1
                    else:
                        created += 1

                report = {"created": created, "pending": pending, "skipped": skipped,
                          "content_updated_existing": updated,
                          "content_written": content_written,
                          "shops": len(shop_ids), "groups": len(group_ids)}
                cur.execute("""
                    UPDATE public.bulk_import_job
                    SET status = 'committed', created_rows = %s, skipped_rows = %s,
                        report = %s::jsonb, finished_at = NOW()
                    WHERE id = %s;
                """, (created + pending, skipped, _json(report), body.job_id))
                log_audit(conn, tenant_id, ctx.user_id, "devices.bulk_import", "device", None,
                          details=report)
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:
            conn.rollback()
            logger.error("bulk commit failed (tenant %s job %s): %s", tenant_id, body.job_id, e)
            raise HTTPException(status_code=500, detail="Bulk import failed; nothing was created")
    return {"job_id": body.job_id, "committed": True, **report}


@router.get("/bulk-devices/jobs/{job_id}")
def bulk_job(job_id: int, ctx: TenantContext = Depends(require_tenant_context)):
    _bulk_mode(ctx)  # content editors poll their own jobs too
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, filename, status, total_rows, created_rows, skipped_rows,
                       error_rows, report, created_at, finished_at
                FROM public.bulk_import_job WHERE id = %s AND tenant_id = %s;
            """, (job_id, ctx.active_tenant_id))
            r = cur.fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Import job not found")
    return {"id": r[0], "filename": r[1], "status": r[2], "total_rows": r[3],
            "created_rows": r[4], "skipped_rows": r[5], "error_rows": r[6],
            "report": r[7], "created_at": r[8].isoformat() if r[8] else None,
            "finished_at": r[9].isoformat() if r[9] else None}


@router.get("/bulk-devices/pending")
def bulk_pending(ctx: TenantContext = Depends(require_tenant_context)):
    """Devices created without a device_id, awaiting an ANDROID_ID (claim)."""
    _require_manage_devices(ctx)
    with pg_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT d.id, d.device_name, d.activation_code, d.resolution,
                       s.shop_name, g.gname, d.created_at
                FROM public.device d
                LEFT JOIN public.device_assignment da ON da.did = d.id
                LEFT JOIN public.shop s ON s.id = da.sid
                LEFT JOIN public."group" g ON g.id = da.gid
                WHERE d.tenant_id = %s AND d.mobile_id LIKE %s
                ORDER BY d.created_at DESC;
            """, (ctx.active_tenant_id, PENDING_PREFIX + "%"))
            items = [{"id": r[0], "device_name": r[1], "code": r[2], "resolution": r[3],
                      "shop_name": r[4], "group_name": r[5],
                      "created_at": r[6].isoformat() if r[6] else None} for r in cur.fetchall()]
    return {"items": items, "count": len(items)}


@router.post("/bulk-devices/claim")
def bulk_claim(body: ClaimIn, ctx: TenantContext = Depends(require_tenant_context)):
    """Bind a real ANDROID_ID to a pending device row. The physical device then
    auto-enrolls on its next GET /device/{id}/online poll."""
    _require_manage_devices(ctx)
    tenant_id = ctx.active_tenant_id
    mobile_id = body.mobile_id.strip()
    if not MOBILE_ID_RE.match(mobile_id) or mobile_id.startswith(PENDING_PREFIX):
        raise HTTPException(status_code=422, detail="Invalid device id")
    with pg_conn() as conn:
        try:
            with conn.cursor() as cur:
                # Reject if this ANDROID_ID is already used in ANY tenant
                # (global heartbeat lookups ignore tenant).
                cur.execute("SELECT tenant_id FROM public.device WHERE mobile_id = %s;", (mobile_id,))
                clash = cur.fetchone()
                if clash:
                    raise HTTPException(status_code=409,
                                        detail="That device id is already registered"
                                               + (" to this company" if clash[0] == tenant_id else " to another company"))
                cur.execute("""
                    UPDATE public.device
                    SET mobile_id = %s, activation_code = NULL, activated_at = NOW(), updated_at = NOW()
                    WHERE id = %s AND tenant_id = %s AND mobile_id LIKE %s
                    RETURNING id, device_name;
                """, (mobile_id, body.device_id, tenant_id, PENDING_PREFIX + "%"))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(status_code=404, detail="Pending device not found (already claimed?)")
                log_audit(conn, tenant_id, ctx.user_id, "device.claim", "device", body.device_id,
                          details={"mobile_id": mobile_id})
            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except Exception as e:
            conn.rollback()
            logger.error("claim failed (tenant %s): %s", tenant_id, e)
            raise HTTPException(status_code=500, detail="Claim failed")
    return {"device_id": body.device_id, "mobile_id": mobile_id, "claimed": True}


# ─────────────────────────── helpers ───────────────────────────

def _json(obj) -> str:
    import json
    return json.dumps(obj)


def _pending_code() -> str:
    # short, unambiguous (no O/0/I/1) human code for the pending device
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    return "".join(secrets.choice(alphabet) for _ in range(6))


def _upsert_named(cur, table: str, name_col: str, names: set, tenant_id: int) -> Dict[str, int]:
    """Idempotent upsert of shops/groups; returns {name: id}. Uses the per-tenant
    unique index so ON CONFLICT ... DO UPDATE ... RETURNING yields the id either way."""
    out: Dict[str, int] = {}
    conflict = "(tenant_id, shop_name)" if name_col == "shop_name" else "(tenant_id, gname)"
    for name in names:
        cur.execute(f"""
            INSERT INTO public.{table} ({name_col}, tenant_id) VALUES (%s, %s)
            ON CONFLICT {conflict} DO UPDATE SET {name_col} = EXCLUDED.{name_col}
            RETURNING id;
        """, (name, tenant_id))
        out[name] = cur.fetchone()[0]
    return out
